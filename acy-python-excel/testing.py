# print("Hello World")
# name = "Win Tun"
# print("Hello",name)

# for i in range(10):
#     print(i)

# for char in "Python":
#     print(char)

from openpyxl import load_workbook


wb = load_workbook("Books1.xlsx")
ws=wb.active
print(ws["A1"])
print(ws["B2"].value)
ws["B2"].value="KauThaung"
ws.save["Books1.xlsx"]